#!/usr/bin/env python3

import psutil
import datetime
import time
import schedule
import openpyxl
import platform
from tqdm import tqdm
from time import sleep
import signal
import sys
import GPUtil as GPU
import os
import subprocess
import pynvml

# Define a simple linear model parameters
CPU_POWER_COEFFICIENT = 0.02  # Coefficient to relate CPU usage to power consumption
BASE_POWER = 10  # Base power consumption
DISK_POWER_COEFFICIENT = 0.001  # Coefficient to relate disk usage to power consumption
MEMORY_POWER_COEFFICIENT = 0.005  # Coefficient to relate memory usage to power consumption
GPU_POWER_COEFFICIENT = 0.1  # Coefficient to relate GPU usage to power consumption

def estimate_power_usage(cpu_usage, disk_usage, memory_usage, gpu_usage):
    # Estimate power consumption based on CPU, disk, memory, and GPU usage
    estimated_power = (
        BASE_POWER
        + (CPU_POWER_COEFFICIENT * cpu_usage)
        + (DISK_POWER_COEFFICIENT * disk_usage)
        + (MEMORY_POWER_COEFFICIENT * memory_usage)
        + (GPU_POWER_COEFFICIENT * gpu_usage)  # Adding GPU usage to power consumption estimation
    )
    return round(estimated_power, 2)

def main():
    pid = int(input("Enter process ID: "))
    
    def exit_program(signal, frame):
        sys.exit(0)
    
    signal.signal(signal.SIGINT, exit_program)
    
    p = psutil.Process(pid)
    process_name = p.name() 
    schedule.every(5).seconds.do(monitor_with_progress, pid, process_name)
  
    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except Exception as e:
        print("An error occurred:", e)

def get_process_disk_usage(pid):
    try:
        process = psutil.Process(pid)
        disk_usage = process.io_counters()
        total_disk_io = disk_usage.write_bytes + disk_usage.read_bytes
        total_disk_space = psutil.disk_usage(os.sep).total
        
        if total_disk_space > 0:
            disk_usage_percentage = (total_disk_io / total_disk_space) * 100
        else:
            disk_usage_percentage = 0
        
        return disk_usage_percentage
    except psutil.NoSuchProcess:
        return 0
    
def gpu_available():
    try:
        subprocess.check_output('nvidia-smi')
        return 1
    except Exception:
        return 0
    
def get_gpu_memory_usage():
    command = "nvidia-smi --query-gpu=memory.used,memory.total --format=csv,nounits"
    memory_info = subprocess.check_output(command.split()).decode('ascii').split('\n')[:-1][1:]

    total_used_memory = 0
    total_memory = 0

    for line in memory_info:
        used_memory, total_memory = map(int, line.split(', '))
        total_used_memory += used_memory

    if total_memory == 0:
        return 0.0  # Avoid division by zero
    else:
        memory_usage_percentage = (total_used_memory / total_memory) * 100
        return memory_usage_percentage

def gpu_see_usage():
    pynvml.nvmlInit()

    handle = pynvml.nvmlDeviceGetHandleByIndex(0)  # Assuming GPU index 0, can iterate for multiple GPUs
    info = pynvml.nvmlDeviceGetMemoryInfo(handle)
    total_memory = info.total
    used_memory = info.used

    if total_memory == 0:
        return 0.0  # Avoid division by zero
    
    memory_usage_percentage = (used_memory / total_memory) * 100
    return memory_usage_percentage

def monitor_with_progress(pid, process_name):
    print("Monitoring process:", process_name)
    sleep(1)
    gpu_active = gpu_available()
    
    # Initialize progress bars
    with tqdm(total=100, desc='% CPU USAGE: ', position=0) as cpubar, \
            tqdm(total=100, desc='% GPU USAGE: ', position=1) as gpubar, \
            tqdm(total=100, desc='% RAM USAGE: ', position=2) as rambar, \
            tqdm(total=100, desc='% DISK USAGE: ', position=3) as diskbar, \
            tqdm(total=100, desc='% POWER CONSUMPTION: ', position=4) as powerconbar:
        try:
            while True:
                current_time = datetime.datetime.now().strftime("%Y%m%d - %H:%M:%S")
                p = psutil.Process(pid)
                process_name = p.name() 
                
                cpu_usage = p.cpu_percent(interval=1) / psutil.cpu_count()
                memory = p.memory_percent()  
                memory_formatted = f"{memory:.2f}"
                
                disk_usage = get_process_disk_usage(pid)
                
                gpu_free_memory = get_gpu_memory_usage()
               
                gpu_usage = gpu_see_usage()
                power_consumption = estimate_power_usage(cpu_usage, disk_usage, memory, gpu_usage)
                
                # Update progress bars
                cpubar.n = cpu_usage
                if gpu_active == 1:
                    gpubar.n = gpu_usage
                rambar.n = memory
                diskbar.n = disk_usage
                powerconbar.n = power_consumption
                
                # Refresh progress bars
                cpubar.refresh()
                if gpu_active == 1:
                    gpubar.refresh()
                rambar.refresh()
                diskbar.refresh()
                powerconbar.refresh()

                path = "Monitor_Result.xlsx"

                try:
                    file = openpyxl.load_workbook(path)
                except FileNotFoundError:
                    file = openpyxl.Workbook()

                sheet = file.active
                if sheet.cell(1, 1).value is None:
                    sheet.cell(column=1, row=1, value="Date & Time")
                    sheet.cell(column=2, row=1, value="Process ID")
                    sheet.cell(column=3, row=1, value="Process Name")
                    sheet.cell(column=4, row=1, value="% - CPU Usage")
                    sheet.cell(column=5, row=1, value="% - GPU Usage")
                    sheet.cell(column=6, row=1, value="% - Memory usage")
                    sheet.cell(column=7, row=1, value="% - Disk usage")
                    sheet.cell(column=8, row=1, value="% - Estimated Power Consumption")
                
                row_data = [current_time, pid, process_name, cpu_usage, gpu_usage, memory_formatted, disk_usage, power_consumption]
                sheet.append(row_data)

                file.save(path)
                
        except KeyboardInterrupt:
            print("\nMonitoring stopped.")
        print("Press Ctrl + C to exit")

if __name__ == "__main__":
    main()